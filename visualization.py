"""
Visualization utilities for task scheduling and assignment.

This module provides functions for visualizing various aspects of the task scheduling
and assignment system, including task dependencies, assignment results, and performance metrics.
"""
import os
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
import networkx as nx
import numpy as np
import pandas as pd
import seaborn as sns
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import math

from models import Task, Employee
from utils.logger import logger


def draw_dependency_graph(
        tasks: List[Task],
        filename: Optional[str] = None,
        layout: str = "spring"
) -> None:
    """
    Draw a graph of task dependencies.

    Args:
        tasks: List of tasks
        filename: File to save the plot (None to display only)
        layout: Graph layout algorithm ('spring', 'dot', 'circular', 'spectral', etc.)
    """
    # Create a directed graph
    G = nx.DiGraph()

    # Add tasks and dependencies to the graph
    task_info = {}  # Store task information for node labels
    for task in tasks:
        # Store task info (skills, hours, priority)
        task_info[task.id] = {
            'skills': ', '.join(task.required_skills),
            'hours': task.estimated_hours,
            'priority': task.priority,
            'due': task.due_date.strftime('%m/%d')
        }

        # Add node with task ID
        G.add_node(task.id)

        # Add edges for dependencies
        for dep_id in task.dependencies:
            G.add_edge(dep_id, task.id)

    # Set up the plot
    plt.figure(figsize=(12, 9))

    # Choose layout based on parameter
    if layout == "dot":
        try:
            from networkx.drawing.nx_agraph import graphviz_layout
            pos = graphviz_layout(G, prog="dot")
        except ImportError:
            logger.warning("Graphviz not available, falling back to spring layout")
            pos = nx.spring_layout(G, seed=42)
    elif layout == "circular":
        pos = nx.circular_layout(G)
    elif layout == "spectral":
        pos = nx.spectral_layout(G)
    else:  # Default to spring layout
        pos = nx.spring_layout(G, seed=42)

    # Set node colors based on priority
    node_colors = []
    for node in G.nodes():
        priority = task_info[node]['priority']
        # Map priority 1-5 to color intensity
        color_intensity = 0.2 + (priority / 5) * 0.8
        node_colors.append((0.1, 0.5, color_intensity))

    # Draw nodes
    nx.draw_networkx_nodes(
        G, pos,
        node_size=1500,
        node_color=node_colors,
        edgecolors="black",
        alpha=0.8
    )

    # Draw edges
    nx.draw_networkx_edges(
        G, pos,
        width=1.5,
        edge_color="gray",
        arrowsize=20,
        arrowstyle="-|>"
    )

    # Draw node labels
    nx.draw_networkx_labels(G, pos, font_size=10, font_weight="bold")

    # Add node annotations (skills, hours, due date)
    for node, (x, y) in pos.items():
        info = task_info[node]
        label = f"Skills: {info['skills']}\nHours: {info['hours']}\nDue: {info['due']}"
        plt.annotate(
            label,
            xy=(x, y),
            xytext=(x, y - 0.1),
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8),
            ha='center',
            va='top',
            fontsize=8
        )

    plt.title("Task Dependency Graph", fontsize=16)
    plt.axis("off")
    plt.tight_layout()

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"Dependency graph saved as {filename}")

    plt.show()


def draw_aco_network(
        task_sequence: List[Task],
        title: str = "ACO Task Sequence",
        filename: Optional[str] = None,
        show_attributes: bool = True
) -> None:
    """
    Draw the task sequence determined by ACO with improved layout to reduce overlapping.
    Uses a structured layout with hair pin bends to prevent edge overlaps.

    Args:
        task_sequence: Sequence of tasks as determined by ACO
        title: Plot title
        filename: File to save the plot (None to display only)
        show_attributes: Whether to show task attributes
    """
    # Create a directed graph for the sequence
    G = nx.DiGraph()

    # Add tasks to the graph with attributes
    task_attrs = {}
    for i, t in enumerate(task_sequence):
        G.add_node(t.id)
        task_attrs[t.id] = {
            'idx': i,
            'skills': ', '.join(t.required_skills),
            'hours': t.estimated_hours,
            'priority': t.priority
        }

    # Add sequence edges
    for i in range(len(task_sequence) - 1):
        G.add_edge(task_sequence[i].id, task_sequence[i + 1].id)

    # Create a larger figure for better spacing
    plt.figure(figsize=(16, 10))

    # Create a custom layout to avoid overlapping
    try:
        # Try using graphviz for better layout if available
        try:
            from networkx.drawing.nx_agraph import graphviz_layout
            # Use specific layout parameters to improve separation and routing
            pos = graphviz_layout(
                G,
                prog="dot",
                args="-Grankdir=LR -Goverlap=false -Gsplines=ortho -Gsep=0.5 -Gnodesep=1.0 -Granksep=2.0"
            )
        except ImportError:
            # Custom grid-based layout if graphviz not available
            logger.warning("Graphviz not available, creating custom grid layout")
            pos = {}

            # Calculate optimal grid dimensions
            nodes = list(G.nodes())
            n_nodes = len(nodes)
            grid_width = int(np.sqrt(n_nodes) * 1.5)  # Make grid wider than tall
            grid_height = int(np.ceil(n_nodes / grid_width))

            # Position nodes in a grid pattern
            for i, node in enumerate(nodes):
                row = i // grid_width
                col = i % grid_width

                # Calculate position with spacing
                x = col * 3.0

                # Stagger rows to reduce direct overlaps
                y = row * 2.0
                if col % 2 == 1:
                    y += 0.5  # Stagger alternating columns

                pos[node] = (x, y)
    except Exception as e:
        logger.warning(f"Error creating layout: {e}. Using basic spring layout.")
        pos = nx.spring_layout(G, k=2.0, iterations=500, seed=42)

    # Set node colors based on priority
    node_colors = []
    for node in G.nodes():
        priority = task_attrs[node]['priority']
        color_intensity = 0.2 + (priority / 5) * 0.8
        node_colors.append((0.1, 0.5, color_intensity))

    # Draw nodes with good size for readability
    nx.draw_networkx_nodes(
        G, pos,
        node_size=2000,
        node_color=node_colors,
        edgecolors="black",
        alpha=0.8
    )

    # Draw edges with hair pin routing to prevent overlaps
    edge_styles = []
    for edge in G.edges():
        # Calculate custom routing parameters for each edge
        source, target = edge
        source_x, source_y = pos[source]
        target_x, target_y = pos[target]

        # Calculate routing parameters - create a hairpin bend
        # If nodes are far apart horizontally, bend more
        dist = abs(target_x - source_x)
        rad = min(0.5, 0.1 + 0.05 * dist)

        # Alternate direction for odd/even edges to prevent overlaps
        index = list(G.edges()).index(edge)
        if index % 2 == 0:
            rad *= -1

        edge_styles.append(rad)

    # Draw edges with custom routing
    for i, edge in enumerate(G.edges()):
        nx.draw_networkx_edges(
            G, pos,
            edgelist=[edge],
            width=1.5,
            arrowsize=15,
            arrowstyle="-|>",
            edge_color="gray",
            connectionstyle=f"arc3,rad={edge_styles[i]}"
        )

    # Draw node labels (IDs only)
    nx.draw_networkx_labels(
        G, pos,
        font_size=10,
        font_weight="bold",
        font_color="white"
    )

    # Draw additional task information with careful positioning
    if show_attributes:
        for node in G.nodes():
            x, y = pos[node]
            attr = task_attrs[node]

            # Display task index in a compact label below node
            plt.text(
                x, y - 0.2,
                f"#{attr['idx'] + 1} | Hrs: {attr['hours']} | Pri: {attr['priority']}",
                ha='center',
                va='center',
                bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="gray", alpha=0.8),
                fontsize=9,
                fontweight='bold'
            )

            # Display skills in a separate location to reduce overlap
            plt.text(
                x, y - 0.4,
                f"Skills: {attr['skills']}",
                ha='center',
                va='center',
                bbox=dict(boxstyle="round,pad=0.2", fc="lightyellow", ec="goldenrod", alpha=0.8),
                fontsize=8
            )

    plt.title(title, fontsize=16, fontweight='bold')
    plt.axis("off")
    plt.tight_layout()

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"ACO network plot saved as {filename}")

    plt.show()


def get_employee_color(employee: Employee) -> str:
    """
    Return a color based on the employee's primary skill.

    Args:
        employee: The employee object

    Returns:
        str: Color name for visualization
    """
    # Skill to color mapping
    SKILL_COLOR_MAP = {
        "Drilling": "#3498db",  # Blue
        "Milling": "#2ecc71",  # Green
        "Lathe": "#e67e22",  # Orange
        "CNC": "#9b59b6",  # Purple
        "NDT": "#e74c3c",  # Red
        "Boring": "#8b4513"  # Brown
    }

    if not employee.skills:
        return "gray"

    # Get the first skill in the set (primary skill)
    primary_skill = next(iter(employee.skills))
    return SKILL_COLOR_MAP.get(primary_skill, "gray")


def plot_gantt_chart_for_assignments(
        approach_name: str,
        tasks: List[Task],
        assigned_dict: Dict[str, Optional[str]],
        employees: List[Employee],
        filename: Optional[str] = None
) -> None:
    """
    Create an improved Gantt chart for task assignments with better alignment of capacity limit bars.

    Args:
        approach_name: Name of the assignment approach
        tasks: List of tasks
        assigned_dict: Dictionary mapping task IDs to employee IDs
        employees: List of employees
        filename: File to save the plot (None to display only)
    """
    # Create mapping for employee indices and employee objects
    employee_index_map = {emp.id: i for i, emp in enumerate(employees)}
    employee_map = {emp.id: emp for emp in employees}

    # Setup the plot
    fig, ax = plt.subplots(figsize=(14, 8))

    # Task start and end times per employee
    employee_schedule = {emp.id: [] for emp in employees}  # List of (task_id, start, end, due_date) per employee
    unassigned_tasks = []

    # Calculate start and end times for each task
    current_time = {emp.id: 0 for emp in employees}  # Current time per employee

    # Compute task start and end times
    for t in tasks:
        assigned_emp_id = assigned_dict.get(t.id)

        if assigned_emp_id is not None and assigned_emp_id in employee_map:
            emp = employee_map[assigned_emp_id]

            # Calculate duration based on employee efficiency
            duration = math.ceil(t.estimated_hours / emp.efficiency_rating)

            # Task starts at the current time for this employee
            start_time = current_time[assigned_emp_id]
            end_time = start_time + duration

            # Add task to employee's schedule
            employee_schedule[assigned_emp_id].append((t.id, start_time, end_time, t.due_date))

            # Update current time for this employee
            current_time[assigned_emp_id] = end_time
        else:
            # Track unassigned tasks
            unassigned_tasks.append(t.id)

    # Define color map for employees based on their skills
    employee_colors = {emp.id: get_employee_color(emp) for emp in employees}

    # Improved bar height for better visibility
    bar_height = 0.6

    # Plot the bars for assigned tasks
    for emp_id, schedule in employee_schedule.items():
        emp_idx = employee_index_map[emp_id]
        emp_color = employee_colors[emp_id]

        for task_id, start, end, due_date in schedule:
            task = next(t for t in tasks if t.id == task_id)

            # Calculate task duration
            duration = end - start

            # Plot the bar
            bar = ax.barh(
                y=emp_idx,
                width=duration,
                left=start,
                height=bar_height,
                align="center",
                color=emp_color,
                edgecolor="black",
                alpha=0.8
            )

            # Calculate due date relative position (in hours from project start)
            project_start = datetime(2023, 10, 1)
            due_hours = (due_date - project_start).total_seconds() / 3600

            # Determine if task is potentially overdue
            is_overdue = end > due_hours

            # Add task ID and due date within the bar
            ax.text(
                start + duration / 2,
                emp_idx,
                f"{task_id}\nDue: {due_date.strftime('%m-%d')}",
                va="center",
                ha="center",
                color="white" if emp_color in ["#3498db", "#9b59b6", "#e74c3c", "#8b4513"] else "black",
                fontsize=8,
                fontweight="bold",
                bbox=dict(boxstyle="round,pad=0.1", fc="none", ec="none", alpha=0.1)
            )

            # Add an indicator if task might be overdue
            if is_overdue:
                ax.text(
                    end,
                    emp_idx,
                    "⚠️",
                    va="center",
                    ha="left",
                    fontsize=10
                )

    # Get overall maximum time for x-axis limits
    all_end_times = [end for schedule in employee_schedule.values() for _, _, end, _ in schedule]
    max_time = max([emp.weekly_available_hours for emp in employees] + all_end_times + [40])  # Default to 40 if empty

    # Add capacity markers for each employee - USING DOTS INSTEAD OF LINES
    for i, emp in enumerate(employees):
        # Calculate exact hours limits
        limit_hours = emp.weekly_available_hours

        # Add a distinct dot marker at the exact limit position
        ax.scatter(
            [limit_hours],  # X position exactly at the hour limit
            [i],  # Y position at the employee's row
            marker='o',  # Circular marker
            s=100,  # Size of marker
            color='red',  # Red color
            edgecolor='black',  # Black edge for contrast
            zorder=10  # Ensure it's drawn on top
        )

        # Add vertical dotted line for reference
        ax.plot(
            [limit_hours, limit_hours],  # Vertical line at exact hour limit
            [i - 0.4, i + 0.4],  # Extending slightly above and below the row
            linestyle=':',  # Dotted line
            color='red',  # Red color
            linewidth=1.5,  # Thin line
            alpha=0.7,  # Slightly transparent
            zorder=5  # Below the dot marker but above other elements
        )

        # Add limit label with exact positioning
        ax.text(
            limit_hours + 0.3,  # Positioned right after the capacity marker
            i,  # At employee's row
            f"Limit: {limit_hours}h",  # Text label
            va="center",  # Vertically centered
            ha="left",  # Left-aligned
            fontsize=9,  # Font size
            color="red",  # Red text
            fontweight="bold",  # Bold text
            bbox=dict(  # Background box
                boxstyle="round,pad=0.2",
                fc="white",  # White background
                ec="red",  # Red edge
                alpha=0.9  # Nearly opaque
            ),
            zorder=15  # On top of everything
        )

    # Plot unassigned tasks at the bottom with improved visibility
    if unassigned_tasks:
        unassigned_row = len(employees)

        # Add a more prominent separator line
        ax.axhline(y=unassigned_row - 0.5, color='black', linestyle='-', linewidth=1.5, alpha=0.8)

        # Add unassigned tasks label with better visibility
        ax.text(
            0,
            unassigned_row,
            f"Unassigned Tasks: {', '.join(unassigned_tasks)}",
            va="center",
            ha="left",
            color="black",
            fontsize=10,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", fc="mistyrose", ec="red", alpha=0.8)
        )

    # Configure plot appearance
    ax.set_yticks(range(len(employees)))
    ax.set_yticklabels([f"{emp.id}: {emp.name} ({', '.join(emp.skills)})" for emp in employees])
    ax.set_xlabel("Time (hours)")
    ax.set_ylabel("Employees (colored by primary skill)")
    ax.set_title(f"Task Assignment Gantt Chart - {approach_name}")

    # Add grid for readability with improved visibility
    ax.grid(axis="x", linestyle="--", alpha=0.4)

    # Add alternating row backgrounds for better readability
    for i in range(len(employees)):
        if i % 2 == 0:  # Every other row
            ax.axhspan(i - 0.5, i + 0.5, color="lightgray", alpha=0.15)

    # Add skill color legend
    skill_patches = []
    for skill, color in {"Drilling": "#3498db", "Milling": "#2ecc71", "Lathe": "#e67e22",
                         "CNC": "#9b59b6", "NDT": "#e74c3c", "Boring": "#8b4513"}.items():
        if any(skill in emp.skills for emp in employees):
            skill_patches.append(mpatches.Patch(color=color, label=skill))

    ax.legend(handles=skill_patches, title="Employee Skills", loc="upper right")

    # Set reasonable x-axis limits
    ax.set_xlim(0, max_time * 1.1)  # Add 10% padding

    plt.tight_layout()

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"Gantt chart saved as {filename}")

    plt.show()


def plot_rewards_across_episodes(
        q_rewards: List[float],
        sarsa_rewards: List[float],
        window_size: int = 50,
        episode_interval: int = 10,
        filename: Optional[str] = None,
) -> None:
    """
    Plot rewards history during training.

    Args:
        q_rewards: Rewards from Q-learning
        sarsa_rewards: Rewards from SARSA
        window_size: Window size for moving average
        episode_interval: Interval for plotting points
        filename: File to save the plot (None to display only)
    """
    plt.figure(figsize=(12, 6))

    # Filter episodes for plotting (to avoid overcrowding)
    filtered_episodes = range(0, len(q_rewards), episode_interval)
    q_rewards_filtered = [q_rewards[i] for i in filtered_episodes]
    sarsa_rewards_filtered = [sarsa_rewards[i] for i in filtered_episodes]

    # Compute moving averages
    def moving_avg(data, w_size):
        if w_size <= 1:
            return np.array(data)
        if len(data) < w_size:
            return np.array([])
        return np.convolve(data, np.ones(w_size) / w_size, mode="valid")

    q_moving_avg = moving_avg(q_rewards, window_size)
    sarsa_moving_avg = moving_avg(sarsa_rewards, window_size)

    # Plot raw data with transparency
    plt.plot(
        filtered_episodes,
        q_rewards_filtered,
        label="Q-learning (raw)",
        color="blue",
        alpha=0.3,
        marker=".",
        markersize=3,
    )
    plt.plot(
        filtered_episodes,
        sarsa_rewards_filtered,
        label="SARSA (raw)",
        color="orange",
        alpha=0.3,
        marker=".",
        markersize=3,
    )

    # Plot moving averages
    if len(q_moving_avg) > 0:
        moving_avg_episodes = range(window_size - 1, len(q_rewards))
        plt.plot(
            moving_avg_episodes,
            q_moving_avg,
            label=f"Q-learning {window_size}-ep MA",
            color="blue",
            linewidth=2,
        )

    if len(sarsa_moving_avg) > 0:
        moving_avg_episodes = range(window_size - 1, len(sarsa_rewards))
        plt.plot(
            moving_avg_episodes,
            sarsa_moving_avg,
            label=f"SARSA {window_size}-ep MA",
            color="orange",
            linewidth=2,
        )

    # Configure plot
    plt.title(
        f"Rewards During Training (Every {episode_interval} Episodes)"
    )
    plt.xlabel("Episode")
    plt.ylabel("Total Reward")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"Rewards plot saved as {filename}")

    plt.show()


def plot_epsilon_decay(
        eps_q: List[float], eps_sarsa: List[float], filename: Optional[str] = None
) -> None:
    """
    Plot epsilon decay during training.

    Args:
        eps_q: Epsilon values from Q-learning
        eps_sarsa: Epsilon values from SARSA
        filename: File to save the plot (None to display only)
    """
    plt.figure(figsize=(10, 6))

    # Plot epsilon values
    plt.plot(eps_q, label="Q-learning Epsilon", marker="o", markersize=3)
    plt.plot(eps_sarsa, label="SARSA Epsilon", marker="o", markersize=3)

    # Configure plot
    plt.xlabel("Episode")
    plt.ylabel("Epsilon (Exploration Rate)")
    plt.title("Exploration Rate (Epsilon) Decay During Training")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"Epsilon decay plot saved as {filename}")

    plt.show()


def plot_metric_bar_chart(
        metric_name: str, data: Dict[str, float], filename: Optional[str] = None
) -> None:
    """
    Plot a bar chart for a specific metric across different methods.

    Args:
        metric_name: Name of the metric
        data: Dictionary mapping method names to metric values
        filename: File to save the plot (None to display only)
    """
    plt.figure(figsize=(10, 6))

    # Prepare data
    methods = list(data.keys())
    values = [data[m] for m in methods]

    # Create bars with sequential color palette
    colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(methods)))
    bars = plt.bar(methods, values, color=colors)

    # Prepare ylabel based on metric name
    ylabel = metric_name.replace("_", " ").title()
    if metric_name == "workload_balance_ratio":
        ylabel += " (Lower is better)"

    # Configure plot
    plt.xlabel("Assignment Method")
    plt.ylabel(ylabel)
    plt.title(f"{ylabel} by Method")

    # Add value labels above bars
    for bar in bars:
        height = bar.get_height()
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            height,
            f"{height:.3f}",
            ha="center",
            va="bottom",
            fontsize=9,
        )

    # Add grid for readability
    plt.grid(axis="y", linestyle="--", alpha=0.3)
    plt.tight_layout()

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"Plot for {metric_name} saved as {filename}")

    plt.show()


def create_comparison_dashboard(
        methods_data: Dict[str, Dict[str, float]],
        filename: Optional[str] = None,
) -> None:
    """
    Create a simplified dashboard comparing different assignment methods.

    Args:
        methods_data: Dictionary mapping method names to metric dictionaries
        filename: File to save the dashboard (None to display only)
    """
    if not methods_data:
        logger.warning("No data provided for dashboard.")
        return

    # Set a unified style for the plots
    plt.style.use('seaborn-v0_8-whitegrid')

    # Create a figure with subplots - simplified to just 2 plots and a table
    fig, axs = plt.subplots(2, 1, figsize=(14, 10))
    fig.suptitle('Assignment Methods Comparison Dashboard', fontsize=16)

    # Extract methods and metrics
    methods = list(methods_data.keys())

    # Filter out priority ratio since it's not needed
    filtered_metrics = {}
    for method, metrics in methods_data.items():
        filtered_metrics[method] = {k: v for k, v in metrics.items() if k != 'assigned_priority_ratio'}

    metrics = set()
    for m_data in filtered_metrics.values():
        metrics.update(m_data.keys())

    metrics = sorted(list(metrics))

    # 1. Bar chart for workload balance (lower is better)
    if 'workload_balance_ratio' in metrics:
        ax = axs[0]
        balance_data = {m: filtered_metrics[m].get('workload_balance_ratio', 0) for m in methods}
        colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(methods)))
        bars = ax.bar(balance_data.keys(), balance_data.values(), color=colors)

        # Add data labels
        for bar in bars:
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                height,
                f"{height:.3f}",
                ha="center",
                va="bottom",
                fontsize=9,
            )

        ax.set_title('Workload Balance Ratio (Lower is Better)')
        ax.set_xlabel('Assignment Method')
        ax.set_ylabel('Std Dev / Mean Workload')
        ax.grid(axis="y", linestyle="--", alpha=0.3)

    # 2. Bar chart for resource utilization (higher is better)
    if 'resource_utilization' in metrics:
        ax = axs[1]
        util_data = {m: filtered_metrics[m].get('resource_utilization', 0) for m in methods}
        colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(methods)))
        bars = ax.bar(util_data.keys(), util_data.values(), color=colors)

        # Add data labels
        for bar in bars:
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                height,
                f"{height:.1f}%",
                ha="center",
                va="bottom",
                fontsize=9,
            )

        ax.set_title('Resource Utilization (Higher is Better)')
        ax.set_xlabel('Assignment Method')
        ax.set_ylabel('Utilization Percentage')
        ax.grid(axis="y", linestyle="--", alpha=0.3)

    # Adjust layout
    plt.tight_layout(rect=[0, 0, 1, 0.95])

    # Save or display
    if filename:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        logger.info(f"Comparison dashboard saved as {filename}")

    plt.show()

    # Create a separate figure for the summary table
    plt.figure(figsize=(10, 5))
    ax = plt.gca()
    ax.axis('tight')
    ax.axis('off')

    # Prepare table data
    table_data = []
    for metric in metrics:
        if metric != 'assigned_priority_ratio':  # Skip priority ratio
            row = [metric.replace('_', ' ').title()]

            for method in methods:
                if metric in filtered_metrics[method]:
                    # Format based on metric type
                    if metric == 'workload_balance_ratio':
                        row.append(f"{filtered_metrics[method][metric]:.3f}")
                    elif 'ratio' in metric or 'rate' in metric or 'utilization' in metric or 'coverage' in metric:
                        row.append(f"{filtered_metrics[method][metric]:.1f}%")
                    else:
                        row.append(f"{filtered_metrics[method][metric]:.2f}")
                else:
                    row.append('N/A')

            table_data.append(row)

    # Create the table
    table = ax.table(
        cellText=table_data,
        colLabels=['Metric'] + methods,
        loc='center',
        cellLoc='center',
    )

    # Configure table appearance
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.scale(1, 1.5)

    plt.title("Performance Metrics Summary", fontsize=14)
    plt.tight_layout()

    # Save summary table
    if filename:
        table_filename = filename.replace('.png', '_table.png')
        plt.savefig(table_filename, dpi=150, bbox_inches="tight")
        logger.info(f"Summary table saved as {table_filename}")

    plt.show()


def export_to_excel(
        filename: str,
        employees_q: List[Employee],
        employees_sarsa: List[Employee],
        employees_greedy: List[Employee],
        tasks: List[Task],
        assigned_q_before: Dict[str, Optional[str]],
        assigned_sarsa_before: Dict[str, Optional[str]],
        assigned_greedy_before: Dict[str, Optional[str]],
        assigned_q_after: Dict[str, Optional[str]],
        assigned_sarsa_after: Dict[str, Optional[str]],
) -> bool:
    """
    Export assignment results to Excel for analysis.

    Args:
        filename: File to save the Excel spreadsheet
        employees_q: Employees for Q-learning
        employees_sarsa: Employees for SARSA
        employees_greedy: Employees for Greedy
        tasks: List of tasks
        assigned_q_before: Q-learning assignments before refinement
        assigned_sarsa_before: SARSA assignments before refinement
        assigned_greedy_before: Greedy assignments
        assigned_q_after: Q-learning assignments after refinement
        assigned_sarsa_after: SARSA assignments after refinement

    Returns:
        bool: True if export successful
    """
    # Create a new workbook
    wb = Workbook()

    # Define some styles for better readability
    header_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    # Employee sheet
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["Employee ID", "Name", "Skills", "WeeklyHrs", "EffRating"])

    # Apply styles to header row
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add employee data
    for emp in employees_q:
        ws1.append([
            emp.id,
            emp.name,
            ",".join(emp.skills),
            emp.weekly_available_hours,
            emp.efficiency_rating,
        ])

    # Tasks sheet
    ws2 = wb.create_sheet("Tasks")
    ws2.append(["TaskID", "Name", "ReqSkills", "EstHrs", "DueDate", "Priority", "Deps"])

    # Apply styles to header row
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add task data
    for t in tasks:
        ws2.append([
            t.id,
            t.name,
            ",".join(t.required_skills),
            t.estimated_hours,
            t.due_date.strftime("%Y-%m-%d"),
            t.priority,
            ",".join(list(t.dependencies)),
        ])

    # Create employee dictionaries for lookup
    emp_q_dict = {emp.id: emp for emp in employees_q}
    emp_sarsa_dict = {emp.id: emp for emp in employees_sarsa}
    emp_greedy_dict = {emp.id: emp for emp in employees_greedy}

    # Assignments sheet
    ws3 = wb.create_sheet("Assignments")
    ws3.append([
        "TaskID",
        "Q-learning Before",
        "Q-learning Validity Before",
        "SARSA Before",
        "SARSA Validity Before",
        "Greedy Before",
        "Greedy Validity Before",
        "Optimized Q-learning After",
        "Optimized Q-learning Validity After",
        "Optimized SARSA After",
        "Optimized SARSA Validity After",
    ])

    # Apply styles to header row
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Validation function
    def validate_assignment(assigned_emp_id, emp_dict, task):
        if assigned_emp_id is None:
            return "N/A"
        emp = emp_dict.get(assigned_emp_id)
        if emp is None:
            return "Unknown"
        if not task.required_skills.intersection(emp.skills):
            return "No"
        needed_hours = math.ceil(task.estimated_hours / emp.efficiency_rating)
        if emp.assigned_hours > emp.weekly_available_hours:
            return "Overload"
        return "Yes"

    # Sort tasks for consistent output
    sorted_tasks = sorted(tasks, key=lambda x: x.id)

    # Add task assignments
    for t in sorted_tasks:
        q_before = assigned_q_before.get(t.id, "")
        sarsa_before = assigned_sarsa_before.get(t.id, "")
        greedy_before = assigned_greedy_before.get(t.id, "")
        q_after = assigned_q_after.get(t.id, "")
        sarsa_after = assigned_sarsa_after.get(t.id, "")

        q_before_valid = validate_assignment(q_before, emp_q_dict, t) if q_before else ""
        sarsa_before_valid = validate_assignment(sarsa_before, emp_sarsa_dict, t) if sarsa_before else ""
        greedy_before_valid = validate_assignment(greedy_before, emp_greedy_dict, t) if greedy_before else ""
        q_after_valid = validate_assignment(q_after, emp_q_dict, t) if q_after else ""
        sarsa_after_valid = validate_assignment(sarsa_after, emp_sarsa_dict, t) if sarsa_after else ""

        ws3.append([
            t.id,
            q_before,
            q_before_valid,
            sarsa_before,
            sarsa_before_valid,
            greedy_before,
            greedy_before_valid,
            q_after,
            q_after_valid,
            sarsa_after,
            sarsa_after_valid,
        ])

    # Workload sheet
    ws4 = wb.create_sheet("Workload")
    ws4.append([
        "EmployeeID",
        "Q(Hrs) Before",
        "SARSA(Hrs) Before",
        "Greedy(Hrs) Before",
        "Optimized Q(Hrs) After",
        "Optimized SARSA(Hrs) After",
        "WeeklyAvail",
        "Utilization %"
    ])

    # Apply styles to header row
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Function to compute workload
    def compute_workload(
            employees: List[Employee], assignments: Dict[str, Optional[str]]
    ) -> Dict[str, int]:
        workload = {emp.id: 0 for emp in employees}
        for t in tasks:
            employee_id = assignments.get(t.id)
            if employee_id is not None:
                emp = next((emp for emp in employees if emp.id == employee_id), None)
                if emp:
                    needed = math.ceil(t.estimated_hours / emp.efficiency_rating)
                    workload[employee_id] += needed
        return workload

    # Compute workload for each assignment method
    workload_q_before = compute_workload(employees_q, assigned_q_before)
    workload_sarsa_before = compute_workload(employees_sarsa, assigned_sarsa_before)
    workload_greedy_before = compute_workload(employees_greedy, assigned_greedy_before)
    workload_q_after = compute_workload(employees_q, assigned_q_after)
    workload_sarsa_after = compute_workload(employees_sarsa, assigned_sarsa_after)

    # Get all employee IDs and sort for consistent output
    employee_ids = sorted({emp.id for emp in employees_q})

    # Add employee workload data
    for eid in employee_ids:
        # Find employee information
        emp = next((e for e in employees_q if e.id == eid), None)
        if not emp:
            continue

        # Calculate utilization percentage
        weekly_avail = emp.weekly_available_hours
        q_after_hours = workload_q_after.get(eid, 0)
        utilization_pct = round((q_after_hours / weekly_avail * 100) if weekly_avail > 0 else 0, 1)

        # Add row to worksheet
        ws4.append([
            eid,
            workload_q_before.get(eid, 0),
            workload_sarsa_before.get(eid, 0),
            workload_greedy_before.get(eid, 0),
            workload_q_after.get(eid, 0),
            workload_sarsa_after.get(eid, 0),
            weekly_avail,
            f"{utilization_pct}%"
        ])

    # Add summary statistics row
    q_before_total = sum(workload_q_before.values())
    sarsa_before_total = sum(workload_sarsa_before.values())
    greedy_before_total = sum(workload_greedy_before.values())
    q_after_total = sum(workload_q_after.values())
    sarsa_after_total = sum(workload_sarsa_after.values())
    total_avail = sum(emp.weekly_available_hours for emp in employees_q)

    ws4.append([
        "TOTAL",
        q_before_total,
        sarsa_before_total,
        greedy_before_total,
        q_after_total,
        sarsa_after_total,
        total_avail,
        f"{round(q_after_total / total_avail * 100, 1)}%"
    ])

    # Apply bold to the total row
    for cell in ws4[len(employee_ids) + 2]:
        cell.font = Font(bold=True)

    # Summary sheet
    ws5 = wb.create_sheet("Summary")
    ws5.append(["Metric", "Q-learning Before", "SARSA Before", "Greedy Before",
                "Q-learning After", "SARSA After"])

    # Apply styles to header row
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Count assigned tasks
    q_before_assigned = sum(1 for v in assigned_q_before.values() if v is not None)
    sarsa_before_assigned = sum(1 for v in assigned_sarsa_before.values() if v is not None)
    greedy_before_assigned = sum(1 for v in assigned_greedy_before.values() if v is not None)
    q_after_assigned = sum(1 for v in assigned_q_after.values() if v is not None)
    sarsa_after_assigned = sum(1 for v in assigned_sarsa_after.values() if v is not None)

    total_tasks = len(tasks)

    # Add task assignment counts
    ws5.append([
        "Tasks Assigned",
        f"{q_before_assigned}/{total_tasks} ({round(q_before_assigned / total_tasks * 100, 1)}%)",
        f"{sarsa_before_assigned}/{total_tasks} ({round(sarsa_before_assigned / total_tasks * 100, 1)}%)",
        f"{greedy_before_assigned}/{total_tasks} ({round(greedy_before_assigned / total_tasks * 100, 1)}%)",
        f"{q_after_assigned}/{total_tasks} ({round(q_after_assigned / total_tasks * 100, 1)}%)",
        f"{sarsa_after_assigned}/{total_tasks} ({round(sarsa_after_assigned / total_tasks * 100, 1)}%)"
    ])

    # Calculate workload balance (standard deviation)
    q_before_std = np.std(list(workload_q_before.values())) if workload_q_before else 0
    sarsa_before_std = np.std(list(workload_sarsa_before.values())) if workload_sarsa_before else 0
    greedy_before_std = np.std(list(workload_greedy_before.values())) if workload_greedy_before else 0
    q_after_std = np.std(list(workload_q_after.values())) if workload_q_after else 0
    sarsa_after_std = np.std(list(workload_sarsa_after.values())) if workload_sarsa_after else 0

    # Add workload balance row
    ws5.append([
        "Workload Balance (std dev)",
        f"{q_before_std:.2f} hours",
        f"{sarsa_before_std:.2f} hours",
        f"{greedy_before_std:.2f} hours",
        f"{q_after_std:.2f} hours",
        f"{sarsa_after_std:.2f} hours"
    ])

    # Calculate resource utilization
    q_before_util = q_before_total / total_avail * 100 if total_avail > 0 else 0
    sarsa_before_util = sarsa_before_total / total_avail * 100 if total_avail > 0 else 0
    greedy_before_util = greedy_before_total / total_avail * 100 if total_avail > 0 else 0
    q_after_util = q_after_total / total_avail * 100 if total_avail > 0 else 0
    sarsa_after_util = sarsa_after_total / total_avail * 100 if total_avail > 0 else 0

    # Add resource utilization row
    ws5.append([
        "Resource Utilization",
        f"{q_before_util:.1f}%",
        f"{sarsa_before_util:.1f}%",
        f"{greedy_before_util:.1f}%",
        f"{q_after_util:.1f}%",
        f"{sarsa_after_util:.1f}%"
    ])

    # Adjust column widths for better readability
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    length = len(str(val))
                    if length > max_len:
                        max_len = length
            sheet.column_dimensions[col_letter].width = max_len + 2

    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Save the workbook
    try:
        wb.save(filename)
        logger.info(f"Excel report saved as '{filename}'")
        return True
    except Exception as e:
        logger.error(f"Error saving Excel report: {e}")
        return False